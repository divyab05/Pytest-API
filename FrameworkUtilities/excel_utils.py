import collections
import os
import openpyxl


# Read Data from Excel based on file path and sheet_name #


def read_from_excel(file_path, sheet_name):
    if os.path.exists(file_path) is True:
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            ws = wb[sheet_name]
            result = []
            first_row = []
            row_index = 0

            for row in ws.iter_rows():
                list_item = collections.OrderedDict()
                if row_index == 0:
                    for cell in row:
                        first_row.append(str(cell.value))
                    row_index = row_index + 1
                else:
                    column_index = 0
                    for cell in row:
                        list_item[first_row[column_index]] = str(cell.value)
                        column_index = column_index + 1
                    row_index = row_index + 1
                    result.append(list_item)

            row_cnt = len(result)

            if row_cnt > 0:
                col_cnt = len(result[0])
            else:
                col_cnt = 0
            final_results = {"Rowcnt": row_cnt, "ColCnt": col_cnt, "Records": result}
            return final_results

        except Exception as e:
            print("Exception Message : {arg}".format(arg=e))
    else:
        print("File does not exists in path: {0}".format(file_path))


'''
Read data from Excel based on a particular cell value and required three arguments
file_path: path of the excel file
sheet_name: sheet name from the file_path provided
col_name: column value 
'''


def read_excel_based_on_col_value(file_path, sheet_name, col_name):
    if os.path.exists(file_path) is True:
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            ws = wb[sheet_name]
            result = []
            first_row = []
            row_index = 0

            for row in ws.iter_rows():
                list_item = collections.OrderedDict()
                if row_index == 0:
                    for cell in row:
                        first_row.append(str(cell.value))
                    row_index = row_index + 1
                else:
                    column_index = 0

                    for cell in row:
                        if str(cell.value).strip() == str(col_name).strip():
                            for i in range(1, ws.max_column + 1):
                                list_item[first_row[column_index]] = ws.cell(row=cell.row, column=i).value
                                column_index = column_index + 1
                            row_index = row_index + 1
                            result.append(list_item)

            return result
        except Exception as e:
            print("Exception Message : {arg}".format(arg=e))

    else:
        print("File does not exists in path: {0}".format(file_path))
